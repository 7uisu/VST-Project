from django.shortcuts import render, redirect, get_object_or_404
from .decorators import login_required, admin_required
from django.contrib.auth.forms import PasswordResetForm, SetPasswordForm
from django.contrib.auth.tokens import default_token_generator
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth import get_user_model
from django.core.mail import send_mail
from django.contrib.sites.shortcuts import get_current_site
from django.contrib import messages

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import csv

from django.contrib.auth import login, authenticate
from django.urls import reverse
from django.forms import modelformset_factory
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LogoutView
from django.views import View
from .models import *
from .forms import *
from django.http import HttpResponse
import logging
from collections import OrderedDict

# Logger setup
logger = logging.getLogger(__name__)

User = get_user_model()

def home_page(request):
    carousel_images = UpcomingEventsCarouselImage.objects.all()
    return render (request, 'main/pages/home.html', {'carousel_images': carousel_images})

def services_page(request):
    return render (request, 'main/pages/services.html')

def events_page(request):
    events = Event.objects.all()
    return render(request, 'main/pages/events.html', {'events': events})

def event_detail_view(request, event_id):
    event = get_object_or_404(Event, pk=event_id)
    return render(request, 'main/pages/event_detail.html', {'event': event})

def contacts_page(request):
    return render (request, 'main/pages/contacts.html')


def admin_register(request):
    if request.method == 'POST':
        form = AdminRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False  # Inactivate account until verified
            user.verification_code = get_random_string(length=6, allowed_chars='0123456789')
            user.verification_code_expiration = timezone.now() + timedelta(hours=1)
            user.save()
            send_verification_email(user)
            request.session['pending_user_id'] = user.id
            return redirect('verify-email')
    else:
        form = AdminRegistrationForm()
    return render(request, 'main/pages/registration-pages/admin_register.html', {'form': form})

def send_verification_email(user):
    verification_code = user.verification_code
    user.save()
    
    send_mail(
        'Your Verification Code',
        f'Your verification code is {verification_code}.',
        'noreply@vstlivestockcorporation.net',
        [user.email],
        fail_silently=False,
    )

def verify_email(request):
    user_id = request.session.get('pending_user_id')
    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == 'POST':
        form = VerificationForm(request.POST)
        if 'resend_code' in request.POST:
            user.verification_code = get_random_string(length=6, allowed_chars='0123456789')
            user.verification_code_expiration = timezone.now() + timedelta(hours=1)
            user.save()
            send_verification_email(user)
            return render(request, 'main/pages/registration-pages/verify_email.html', {'form': form, 'message': 'New verification code sent to your email.'})
        elif 'cancel' in request.POST:
            user.delete()
            return redirect('home')
        else:
            if form.is_valid():
                if user.verification_code == form.cleaned_data['code'] and user.verification_code_expiration > timezone.now():
                    user.is_active = True
                    user.save()
                    login(request, user)
                    return redirect('admin-login')
                else:
                    form.add_error('code', 'Invalid or expired verification code.')
    else:
        form = VerificationForm()
    return render(request, 'main/pages/registration-pages/verify_email.html', {'form': form})



def user_register(request):
    next_url = request.GET.get('next', reverse('user-login'))
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            return redirect(next_url)
    else:
        form = UserRegistrationForm()
    return render(request, 'main/pages/registration-pages/user_register.html', {'form': form, 'next': next_url})


# Admin Login View
def admin_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')  # Using email instead of username
            password = form.cleaned_data.get('password')
            user = authenticate(request, email=email, password=password)
            if user is not None and user.is_admin_user:
                login(request, user)
                return redirect('admin-panel')
            else:
                form.add_error(None, "Invalid credentials or not an admin user")
    else:
        form = AuthenticationForm()
    return render(request, 'main/pages/login-pages/admin_login.html', {'form': form})

# User Login View
def user_login(request):
    next_url = request.GET.get('next', reverse('feedback'))
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            email = form.cleaned_data.get('username')  # Using email instead of username
            password = form.cleaned_data.get('password')
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                return redirect(next_url)
            else:
                form.add_error(None, "Invalid credentials")
    else:
        form = AuthenticationForm()
    return render(request, 'main/pages/login-pages/user_login.html', {'form': form, 'next': next_url})

def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            email = password_reset_form.cleaned_data['email']
            associated_users = User.objects.filter(email=email)
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Requested"
                    email_template_name = "main/pages/password-reset/password_reset_email.txt"
                    c = {
                        "email": user.email,
                        'domain': get_current_site(request).domain,
                        'site_name': 'Your site name',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                    }
                    email = render_to_string(email_template_name, c)
                    send_mail(subject, email, 'noreply@vstlivestockcorporation.net', [user.email], fail_silently=False)
            return redirect("password-reset-done")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="main/pages/password-reset/password_reset.html", context={"password_reset_form": password_reset_form})

def password_reset_confirm(request, uidb64=None, token=None):
    UserModel = get_user_model()
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = UserModel.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, UserModel.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Your password has been successfully reset.')
                return redirect('password-reset-complete')
        else:
            form = SetPasswordForm(user)
        return render(request, 'main/pages/password-reset/password_reset_confirm.html', {'form': form})
    else:
        messages.error(request, 'The password reset link is invalid, possibly because it has already been used. Please request a new password reset.')
        return redirect('password-reset')


logger = logging.getLogger(__name__)

# Admin Panel View
@admin_required(login_url='admin-login')
def admin_panel(request):
    questions = Question.objects.all()
    all_responses = Response.objects.select_related('user', 'question').all()
    
    user_responses = {}
    for response in all_responses:
        user = response.user
        if user not in user_responses:
            user_responses[user] = {}
        user_responses[user][response.question.id] = response

    total_users_responded = len(user_responses)

    recent_responses = all_responses.order_by('-submitted_at')
    recent_users = OrderedDict()
    for response in recent_responses:
        if response.user not in recent_users:
            recent_users[response.user] = response
        if len(recent_users) >= 1:
            break
    recent_responses = list(recent_users.values())

    for user, user_responses_dict in user_responses.items():
        logger.debug(f'User: {user.email}, Responses: {user_responses_dict}')
    
    context = {
        'user': request.user,
        'questions': questions,
        'total_responses': total_users_responded,
        'recent_responses': recent_responses,
        'user_responses': user_responses,
    }
    
    return render(request, 'main/pages/admin-pages/admin_panel.html', context)



@admin_required(login_url='admin-login')
def admin_panel_carousel_images(request):
    carousel_images = UpcomingEventsCarouselImage.objects.all()
    total_images = carousel_images.count()
    recent_caption = carousel_images.latest('created_at').caption if total_images > 0 else "No images available"

    context = {
        'carousel_images': carousel_images,
        'total_images': total_images,
        'recent_caption': recent_caption,
    }

    return render(request, 'main/pages/admin-pages/admin_panel_carousel_images.html', context)

@admin_required(login_url='admin-login')
def admin_panel_more_events(request):
    events = Event.objects.all()
    total_events = events.count()
    recent_event = events.latest('created_at') if total_events > 0 else None

    context = {
        'events': events,
        'total_events': total_events,
        'recent_event': recent_event,
    }

    return render(request, 'main/pages/admin-pages/admin_panel_more_events.html', context)

class CustomLogoutView(LogoutView):
    http_method_names = ['get', 'post']
    next_page = 'home'  # Redirect to the home page after logout

class FeedbackView(LoginRequiredMixin, View):
    login_url = 'user-login'
    redirect_field_name = 'next'

    def get(self, request):
        return render(request, 'main/pages/feedback.html')

# For More Events
@admin_required(login_url='admin-login')
def add_event(request):
    ImageFormSet = modelformset_factory(EventImage, form=EventImageForm, extra=5)  # Allow up to 5 images

    if request.method == 'POST':
        event_form = EventForm(request.POST, request.FILES)
        formset = ImageFormSet(request.POST, request.FILES, queryset=EventImage.objects.none())

        if event_form.is_valid() and formset.is_valid():
            event = event_form.save()
            for form in formset.cleaned_data:
                if form:
                    image = form['image']
                    photo = EventImage(event=event, image=image)
                    photo.save()
            return redirect('event-detail', event_id=event.id)  # Correct the redirect
    else:
        event_form = EventForm()
        formset = ImageFormSet(queryset=EventImage.objects.none())

    return render(request, 'main/pages/admin-pages/events-pages/add_event.html', {'event_form': event_form, 'formset': formset})

@admin_required(login_url='admin-login')
def update_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    ImageFormSet = modelformset_factory(EventImage, form=EventImageForm, extra=5, can_delete=True)

    if request.method == 'POST':
        event_form = EventForm(request.POST, request.FILES, instance=event)
        formset = ImageFormSet(request.POST, request.FILES, queryset=event.images.all())

        if event_form.is_valid() and formset.is_valid():
            event = event_form.save()

            # Save the images
            for form in formset:
                if form.cleaned_data.get('image'):
                    image = form.save(commit=False)
                    image.event = event
                    image.save()

            # Handle deletions
            for form in formset.deleted_forms:
                if form.instance.pk:
                    form.instance.delete()

            return redirect('event-detail', event_id=event.id)
    else:
        event_form = EventForm(instance=event)
        formset = ImageFormSet(queryset=event.images.all())

    return render(request, 'main/pages/admin-pages/events-pages/update_event.html', {
        'event_form': event_form,
        'formset': formset
    })

@admin_required(login_url='admin-login')
def delete_event(request, event_id):
    event = get_object_or_404(Event, id=event_id)

    if request.method == 'POST':
        event.delete()
        return redirect('admin-panel-more-events')

    return render(request, 'main/pages/admin-pages/events-pages/delete_event.html', {'event': event})

# For Carousel Image
@admin_required(login_url='admin-login')
def add_carousel_image(request):
    if request.method == 'POST':
        form = UpcomingEventsCarouselForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('home')
    else:
        form = UpcomingEventsCarouselForm()

    return render(request, 'main/pages/admin-pages/upcoming-events-pages/add_upcoming_events.html', {'form': form})

@admin_required(login_url='admin-login')
def update_carousel_image(request, image_id):
    image = get_object_or_404(UpcomingEventsCarouselImage, id=image_id)
    if request.method == 'POST':
        form = UpcomingEventsCarouselForm(request.POST, request.FILES, instance=image)
        if form.is_valid():
            form.save()
            return redirect('admin-panel')
    else:
        form = UpcomingEventsCarouselForm(instance=image)
    return render(request, 'main/pages/admin-pages/upcoming-events-pages/update_upcoming_events.html', {'form': form, 'image': image})

@admin_required(login_url='admin-login')
def delete_carousel_image(request, image_id):
    image = get_object_or_404(UpcomingEventsCarouselImage, id=image_id)
    
    if request.method == 'POST':
        image.delete()
        return redirect('admin-panel-carousel-images')
    
    return render(request, 'main/pages/admin-pages/upcoming-events-pages/delete_upcoming_events.html', {'image': image})

@admin_required(login_url='admin-login')
def add_question(request):
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Question added successfully.')
            return redirect('admin-panel')
    else:
        form = QuestionForm()
    return render(request, 'main/pages/admin-pages/survey-pages/add_question.html', {'form': form})

@admin_required(login_url='admin-login')
def update_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, 'Question updated successfully.')
            return redirect('admin-panel')
    else:
        form = QuestionForm(instance=question)
    return render(request, 'main/pages/admin-pages/survey-pages/update_question.html', {'form': form})

@admin_required(login_url='admin-login')
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully.')
        return redirect('admin-panel')
    return render(request, 'main/pages/admin-pages/survey-pages/delete_question.html', {'question': question})

@admin_required(login_url='admin-login')
def delete_response(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        responses = Response.objects.filter(user=user)
        responses.delete()
        messages.success(request, f'All responses from {user.email} have been deleted successfully.')
        return redirect('admin-panel')
    return render(request, 'main/pages/admin-pages/survey-pages/delete_response.html', {'user': user})

@admin_required(login_url='admin-login')
def delete_all_responses(request):
    if request.method == 'POST':
        Response.objects.all().delete()
        messages.success(request, 'All responses have been deleted successfully.')
        return redirect('admin-panel')
    return render(request, 'main/pages/admin-pages/survey-pages/delete_all_responses.html')

# Export Responses View
@admin_required(login_url='admin-login')
def export_responses(request):
    questions = Question.objects.all()
    responses = Response.objects.select_related('question', 'user').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Responses"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Write header
    headers = ['Email'] + [question.text for question in questions] + ['Submitted At']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border

    # Group responses by user
    user_responses = {}
    for res in responses:
        if res.user not in user_responses:
            user_responses[res.user] = {}
        user_responses[res.user][res.question.id] = res

    # Write data
    for user, responses in user_responses.items():
        row = [
            user.email,
            user.email,
        ]
        for question in questions:
            response_text = responses.get(question.id).text if question.id in responses else "-"
            row.append(response_text)
        submitted_at = responses[list(responses.keys())[0]].submitted_at.astimezone().strftime('%Y-%m-%d %H:%M:%S') if responses else "N/A"
        row.append(submitted_at)
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = alignment
            cell.border = thin_border

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Get current date
    current_date = datetime.now().strftime('%Y-%m-%d')

    # Set the filename with the current date
    filename = f'responses_{current_date}.xlsx'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required(login_url='login')
def feedback(request):
    questions = Question.objects.all()

    if request.method == 'POST':
        for question in questions:
            response_text = request.POST.get(f'response_{question.id}')
            if response_text:
                Response.objects.create(question=question, text=response_text, user=request.user)
        messages.success(request, 'Your answers have been submitted successfully.')
        return redirect('home')

    return render(request, 'main/pages/feedback.html', {'questions': questions})